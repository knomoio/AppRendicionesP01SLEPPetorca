# streamlit_app.py — bloque de firmas estable (sin cortes) y tabla ajustada al ancho útil
import io, os, json
from datetime import date, datetime
from typing import List

import streamlit as st
import pandas as pd
from fpdf import FPDF
import matplotlib.pyplot as plt

# ---------------------------- Config ----------------------------
st.set_page_config(page_title="Rendición de Fondos Fijos P01 – SLEP Petorca", layout="wide")

# ---------------------------- Helpers & State ----------------------------
def init_state():
    if "data" not in st.session_state:
        st.session_state.data = {
            "fondo_inicial": 0.0,
            "gastos": [],
            "meta": {
                "tipo_fondo": "",
                "responsable": "",
                "rut": "",
                "cargo": "",
                "institucion": "",
                "mes_que_rinde": "",
                "fecha_rendicion": "",
                "n_rendicion": "",
                "n_rex": "",
                "fecha_rex": "",
                "observaciones": "",
                "n_egreso_inicial": "",
                "fecha_egreso_inicial": "",
                "saldo_mes_anterior": 0.0,
                "monto_recibido_mes_anterior": 0.0,
                "monto_gasto_transporte": 0.0,
            }
        }
    if "logo_bytes" not in st.session_state:
        st.session_state.logo_bytes = None
        st.session_state.logo_name = None
    if "firmas" not in st.session_state:
        st.session_state.firmas = {
            "encargado": None,
            "directora": None,
            "revisor1": None,
            "jefe_unidad": None,
            "u_finanzas": None,
            "contab_finanzas": None,
            "jefe_adm_fin": None,
        }

def money(x: float) -> str:
    try:
        s = f"{int(round(float(x))):,}".replace(",", ".")
        return f"${s}"
    except Exception:
        return f"${x}"

def parse_float(x) -> float:
    try:
        return float(x)
    except Exception:
        return 0.0

def load_data_from_json(file) -> None:
    try:
        obj = json.load(file)
        fi = parse_float(obj.get("fondo_inicial", 0))
        gastos = obj.get("gastos", [])
        fixed = []
        for g in gastos:
            fixed.append({
                "fecha": g.get("fecha"),
                "monto": parse_float(g.get("monto", 0)),
                "detalle": g.get("detalle") or g.get("descripcion") or "",
                "tipo_doc": g.get("tipo_doc", ""),
                "n_doc": g.get("n_doc", ""),
                "proveedor": g.get("proveedor", ""),
                "nombre_doc": g.get("nombre_doc"),
                "bytes_doc": None,
            })
        meta = st.session_state.data.get("meta", {}).copy()
        meta.update(obj.get("meta", {}))
        st.session_state.data = {"fondo_inicial": fi, "gastos": fixed, "meta": meta}
        st.success("Datos cargados desde JSON.")
    except Exception as e:
        st.error(f"Error al leer JSON: {e}")

def export_data_json() -> bytes:
    data = {
        "fondo_inicial": st.session_state.data["fondo_inicial"],
        "meta": st.session_state.data["meta"],
        "gastos": [
            {
                "fecha": g.get("fecha"),
                "tipo_doc": g.get("tipo_doc",""),
                "n_doc": g.get("n_doc",""),
                "detalle": g.get("detalle",""),
                "proveedor": g.get("proveedor",""),
                "monto": g.get("monto",0),
                "nombre_doc": g.get("nombre_doc")
            }
            for g in st.session_state.data["gastos"]
        ],
    }
    return json.dumps(data, ensure_ascii=False, indent=2).encode("utf-8")

def gastos_df() -> pd.DataFrame:
    df = pd.DataFrame(st.session_state.data["gastos"])
    if not df.empty:
        df = df.assign(
            N=pd.RangeIndex(1, len(df) + 1),
            Fecha=pd.to_datetime(df["fecha"]).dt.date,
            TipoDocumento=df["tipo_doc"].fillna(""),
            NDocumento=df["n_doc"].fillna(""),
            Detalle=df["detalle"].fillna(""),
            Proveedor=df["proveedor"].fillna(""),
            Monto=pd.to_numeric(df["monto"], errors="coerce").fillna(0.0),
        )[["N","Fecha","TipoDocumento","NDocumento","Detalle","Proveedor","Monto"]]
    return df

def totals():
    df = gastos_df()
    if df.empty:
        total = 0.0; cantidad = 0
    else:
        total = float(pd.to_numeric(df["Monto"], errors="coerce").fillna(0).sum())
        cantidad = int(df.shape[0])
    fondo = float(st.session_state.data.get("fondo_inicial") or 0.0)
    saldo = fondo - total
    return fondo, total, saldo, cantidad

def add_gasto(fecha: date, tipo_doc: str, n_doc: str, detalle: str, proveedor: str, monto: float, doc_file):
    nombre_doc = None; bytes_doc = None
    if doc_file is not None:
        nombre_doc = doc_file.name
        bytes_doc = doc_file.read()
    st.session_state.data["gastos"].append({
        "fecha": fecha.strftime("%Y-%m-%d"),
        "tipo_doc": tipo_doc,
        "n_doc": n_doc,
        "detalle": detalle,
        "proveedor": proveedor,
        "monto": float(monto),
        "nombre_doc": nombre_doc, "bytes_doc": bytes_doc
    })

def remove_gastos(indices: List[int]):
    for idx in sorted(indices, reverse=True):
        if 0 <= idx < len(st.session_state.data["gastos"]):
            st.session_state.data["gastos"].pop(idx)

# ---------- PDF helpers (Unicode) ----------
def set_unicode_font(pdf: FPDF) -> bool:
    candidates = [
        "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf",
        "/usr/local/share/fonts/DejaVuSans.ttf",
        "fonts/DejaVuSans.ttf",
    ]
    for path in candidates:
        if os.path.exists(path):
            pdf.add_font("DejaVu", "", path)
            pdf.add_font("DejaVu", "B", path)
            pdf.set_font("DejaVu", size=11)
            return True
    pdf.set_font("Helvetica", size=11)  # fallback sin Unicode
    return False

def safe_text(txt: str, unicode_ok: bool) -> str:
    return txt if unicode_ok else txt.replace("–", "-").encode("latin-1","ignore").decode("latin-1")

def wrap_text_lines(pdf: FPDF, text: str, width_mm: float, pad: float = 1.5) -> list:
    if text is None: return [""]
    s = str(text)
    max_w = max(1.0, width_mm - pad*2)
    lines, line = [], ""

    def fits(t: str) -> bool:
        return pdf.get_string_width(t) <= max_w

    for word in s.split(" "):
        candidate = f"{line} {word}".strip()
        if fits(candidate):
            line = candidate
            continue
        if line:
            lines.append(line); line = ""
        w = word
        while not fits(w) and w:
            lo, hi, best = 1, len(w), 1
            while lo <= hi:
                mid = (lo + hi)//2
                if fits(w[:mid]): best = mid; lo = mid+1
                else: hi = mid-1
            lines.append(w[:best]); w = w[best:]
        line = w
    if line: lines.append(line)
    return lines or [""]

def draw_wrapped_row(pdf: FPDF, values, widths, aligns, line_h=5.2, unicode_ok=True):
    pdf.set_font(pdf.font_family, size=9)
    x0 = pdf.get_x(); y0 = pdf.get_y()
    all_lines = [wrap_text_lines(pdf, safe_text(v if v is not None else "", unicode_ok), w) for v, w in zip(values, widths)]
    max_lines = max((len(ls) for ls in all_lines), default=1)
    row_h = max_lines * line_h
    for txt_lines, w, a in zip(all_lines, widths, aligns):
        x = pdf.get_x(); y = pdf.get_y()
        pdf.multi_cell(w, line_h, "\n".join(txt_lines), border=1, align=a)
        pdf.set_xy(x + w, y)
    pdf.set_xy(x0, y0 + row_h)

# --- ajusta una lista de anchos al ancho total disponible (corrección del corte) ---
def normalize_widths(widths, total):
    s = sum(widths)
    if s <= 0:
        return [total]
    scale = total / s
    scaled = [w * scale for w in widths]
    if len(scaled) > 1:
        scaled[-1] = total - sum(scaled[:-1])  # evita acumulación de redondeo
    return scaled

# ---------- PDF Export ----------
def export_pdf(landscape: bool, logo_mm: int) -> bytes:
    df = gastos_df()
    fondo, total, saldo, cantidad = totals()
    meta = st.session_state.data.get("meta", {})

    pdf = FPDF(orientation="L" if landscape else "P", unit="mm", format="A4")
    pdf.set_auto_page_break(auto=True, margin=12)  # margen inferior claro
    pdf.add_page()
    unicode_ok = set_unicode_font(pdf)

    # Márgenes y ancho útil
    left = 10
    page_w = 297 if landscape else 210
    usable_w = page_w - 2 * left

    # Logo
    if st.session_state.logo_bytes:
        try:
            bio_logo = io.BytesIO(st.session_state.logo_bytes)
            pdf.image(bio_logo, x=left, y=10, w=max(16, min(logo_mm, 40)))
        except Exception:
            pass

    # Título
    pdf.set_xy(left + 45, 10)
    pdf.set_font(pdf.font_family, "B", 12)
    pdf.cell(0, 6, safe_text("SERVICIO LOCAL DE EDUCACIÓN PÚBLICA DE PETORCA", unicode_ok), ln=True)
    pdf.set_x(left + 45); pdf.set_font(pdf.font_family, size=10)
    pdf.cell(0, 5, safe_text("Rendición de Fondos Fijos P01 – SLEP Petorca", unicode_ok), ln=True)
    pdf.set_x(left + 45)
    pdf.cell(0, 5, safe_text(f"Fecha de emisión: {datetime.now():%Y-%m-%d %H:%M}", unicode_ok), ln=True)
    pdf.ln(2)

    def header_row(labels, widths, align="C"):
        pdf.set_font(pdf.font_family, "B", 10)
        for h, w in zip(labels, widths):
            pdf.cell(w, 7, safe_text(h, unicode_ok), 1, 0, align)
        pdf.ln(7)

    def value_row(values, widths, align="L", h=8, font=""):
        pdf.set_font(pdf.font_family, font, 10)
        for v, w in zip(values, widths):
            pdf.cell(w, h, safe_text("" if v is None else str(v), unicode_ok), 1, 0, align)
        pdf.ln(h)

    # Primera grilla (usa ancho útil)
    w1 = [usable_w * i for i in ([0.17, 0.20, 0.18, 0.20, 0.15, 0.10] if landscape else [0.20,0.23,0.19,0.20,0.13,0.05])]
    pdf.set_x(left); header_row(["Tipo de Fondo","Responsable del fondo","Institución","Fecha Rendición","N° Rendición",""], w1)
    vals1 = [meta.get("tipo_fondo",""), meta.get("responsable",""), meta.get("institucion",""),
             meta.get("fecha_rendicion",""), meta.get("n_rendicion",""), ""]
    pdf.set_x(left); value_row(vals1, w1)

    # Fila REX
    w2 = [usable_w * 0.5, usable_w * 0.5]
    pdf.set_x(left); header_row(["N° REX", "Fecha REX"], w2)
    pdf.set_x(left); value_row([meta.get("n_rex",""), meta.get("fecha_rex","")], w2, "L", 8)

    # Tabla de gastos: ahora normalizada al ancho útil (¡se evita el corte!)
    if landscape:
        col_w_raw = [12, 24, 34, 36, 102, 56, 33]
    else:
        col_w_raw = [10, 24, 30, 30, 84, 42, 30]
    col_w = normalize_widths(col_w_raw, usable_w)

    pdf.set_x(left)
    header_row(["N°", "Fecha gasto", "Tipo documento", "N° Documento",
                "Detalle del gasto", "Nombre Proveedor", "Monto"], col_w)

    if df.empty:
        pdf.set_x(left)
        pdf.cell(sum(col_w), 7, safe_text("Sin registros", unicode_ok), 1, 0, "C"); pdf.ln(7)
    else:
        pdf.set_font(pdf.font_family, size=9)
        for _, r in df.iterrows():
            pdf.set_x(left)
            draw_wrapped_row(
                pdf,
                [
                    str(r["N"]),
                    r["Fecha"].strftime("%Y-%m-%d"),
                    r["TipoDocumento"],
                    str(r["NDocumento"]),
                    str(r["Detalle"]),
                    str(r["Proveedor"]),
                    money(float(r["Monto"])),
                ],
                col_w,
                ["C","L","L","L","L","L","R"],
                line_h=5.2,
                unicode_ok=unicode_ok
            )

    # Total
    pdf.set_font(pdf.font_family, "B", 10)
    pdf.set_x(left); pdf.cell(sum(col_w[:-1]), 7, safe_text("Monto Total del Gasto", unicode_ok), 1, 0, "R")
    pdf.cell(col_w[-1], 7, money(total), 1, 0, "R")
    pdf.ln(9)

    # Egreso inicial
    pdf.set_x(left)
    header_row(["N° Egreso Contable Inicial del Fondo", "Fecha de Egreso Inicial del Fondo"], [usable_w*0.5, usable_w*0.5])
    pdf.set_x(left)
    value_row([meta.get("n_egreso_inicial",""), meta.get("fecha_egreso_inicial","")], [usable_w*0.5, usable_w*0.5], "L", 8)

    # Cuadro resumen
    pdf.set_x(left); pdf.set_font(pdf.font_family, "B", 10)
    pdf.cell(usable_w, 7, safe_text("CUADRO RESUMEN RENDICION", unicode_ok), 1, 0, "C"); pdf.ln(7)
    rows = [
        ("Saldo Inicial/Rendición Mes Anterior", parse_float(meta.get("saldo_mes_anterior", 0))),
        ("Monto Recibido Mes anterior", parse_float(meta.get("monto_recibido_mes_anterior", 0))),
        ("Monto Gasto del mes", total),
        ("Monto del gasto del mes Transporte", parse_float(meta.get("monto_gasto_transporte", 0))),
    ]
    saldo_final = rows[0][1] + rows[1][1] - rows[2][1] - rows[3][1]
    w_label, w_val = usable_w * 0.79, usable_w * 0.21
    pdf.set_font(pdf.font_family, size=10)
    for label, val in rows:
        pdf.set_x(left); pdf.cell(w_label, 8, safe_text(label, unicode_ok), 1, 0, "L")
        pdf.cell(w_val, 8, money(float(val)), 1, 0, "R")
        pdf.ln(8)
    pdf.set_font(pdf.font_family, "B", 10)
    pdf.set_x(left); pdf.cell(w_label, 8, safe_text("Saldo Final", unicode_ok), 1, 0, "L")
    pdf.cell(w_val, 8, money(saldo_final), 1, 0, "R")
    pdf.ln(10)

    # ---------------- Firmas (bloque con control de salto) ----------------
    def draw_signature_box(pdf: FPDF, x: float, y: float, w: float, title: str, key: str):
        """Dibuja 1 casilla de firma y devuelve la altura consumida."""
        SIG_IMG_W = 40
        PAD_LR = 10
        line_y = y + 12  # si no hay imagen
        img_bytes = st.session_state.firmas.get(key)

        # Imagen de firma (opcional)
        if img_bytes:
            try:
                img = io.BytesIO(img_bytes)
                img_x = x + (w - SIG_IMG_W) / 2
                pdf.image(img, x=img_x, y=y, w=SIG_IMG_W)
                line_y = y + 22  # baja un poco la línea si hay imagen
            except Exception:
                pass

        # Línea de firma como trazo
        left_x = x + PAD_LR
        right_x = x + w - PAD_LR
        pdf.line(left_x, line_y, right_x, line_y)

        # Título bajo la línea
        pdf.set_xy(x, line_y + 2)
        pdf.set_font(pdf.font_family, size=9)
        pdf.cell(w, 5, safe_text(title, unicode_ok), 0, 0, "C")

        # Altura estándar consumida
        return max(28.0, (line_y - y) + 9)

    def ensure_space_for_block(pdf: FPDF, needed_h: float):
        """Si no hay espacio suficiente, agrega página antes de dibujar."""
        available = (pdf.h - pdf.b_margin) - pdf.get_y()
        if needed_h > available:
            pdf.add_page()

    # Calcula altura total del bloque (rótulo + 3 filas dobles + opcional 4ª fila)
    row_h_est = 32.0
    exclus_h = 9.0
    gap = 6.0
    # Aquí contamos 4 filas (la última para JEFE/(A) ADM Y FIN) para que nunca se corte
    block_h = exclus_h + gap + row_h_est*4 + gap*3

    ensure_space_for_block(pdf, block_h)

    # Rótulo “USO EXCLUSIVO …”
    pdf.set_font(pdf.font_family, "B", 10)
    pdf.set_x(left)
    pdf.cell(usable_w, 7, safe_text("USO EXCLUSIVO SERVICIO LOCAL DE EDUCACIÓN PÚBLICA DE PETORCA", unicode_ok), 1, 0, "C")
    pdf.ln(gap)

    # Dibujo de las filas de firmas (2 por fila)
    x_left = left
    box_w = (usable_w - 10) / 2  # separación 10 mm entre cajas
    x_right = x_left + box_w + 10
    y = pdf.get_y()

    # Fila 1
    h1 = draw_signature_box(pdf, x_left,  y, box_w, "Encargado/a del Fondo", "encargado")
    h2 = draw_signature_box(pdf, x_right, y, box_w, "Director/a Ejecutiva", "directora")
    y += max(h1, h2) + gap

    # Fila 2
    ensure_space_for_block(pdf, row_h_est + gap)
    h1 = draw_signature_box(pdf, x_left,  y, box_w, "Nombre/Firma Revisor/a 1", "revisor1")
    h2 = draw_signature_box(pdf, x_right, y, box_w, "V°B° JEFE UNIDAD", "jefe_unidad")
    y += max(h1, h2) + gap

    # Fila 3
    ensure_space_for_block(pdf, row_h_est + gap)
    h1 = draw_signature_box(pdf, x_left,  y, box_w, "V°B° UNIDAD DE FINANZAS", "u_finanzas")
    h2 = draw_signature_box(pdf, x_right, y, box_w, "CONTABILIDAD Y FINANZAS", "contab_finanzas")
    y += max(h1, h2) + gap

    # Fila 4
    ensure_space_for_block(pdf, row_h_est)
    _ = draw_signature_box(pdf, x_left, y, box_w, "V°B° JEFE/(A) ADMINISTRACIÓN Y FINANZAS", "jefe_adm_fin")
    pdf.set_y(y + row_h_est)

    out = io.BytesIO()
    pdf.output(out)
    return out.getvalue()

# ---------- Excel Export ----------
def export_excel(logo_px: int) -> bytes:
    from openpyxl import Workbook
    from openpyxl.utils.dataframe import dataframe_to_rows
    from openpyxl.styles import Alignment, Font, Border, Side
    from openpyxl.worksheet.page import PageMargins
    from openpyxl.drawing.image import Image as XLImage

    thin = Side(style="thin", color="000000")
    border = Border(top=thin, left=thin, right=thin, bottom=thin)

    def set_border_range(ws, rng: str):
        for row in ws[rng]:
            for c in row:
                c.border = border

    wb = Workbook()
    # ---------- Hoja Gastos ----------
    ws = wb.active
    ws.title = "Gastos"
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_margins = PageMargins(left=0.3, right=0.3, top=0.5, bottom=0.5)

    # Logo
    if st.session_state.logo_bytes:
        try:
            lf = "/tmp/logo_tmp.png"
            with open(lf, "wb") as f: f.write(st.session_state.logo_bytes)
            img = XLImage(lf)
            img.width = max(80, min(logo_px, 220))
            img.height = int(img.width * 0.35)
            img.anchor = "A1"
            ws.add_image(img)
        except Exception:
            pass

    # Columnas
    headers = ["N°","Fecha del gasto","Tipo documento","N° Documento","Detalle del gasto","Nombre Proveedor","Monto"]
    widths = [6,14,18,20,50,28,14]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[chr(64+i)].width = w

    def merge_set(sheet, cell1, cell2, value="", bold=False, align="center"):
        sheet.merge_cells(f"{cell1}:{cell2}")
        c = sheet[cell1]
        c.value = value
        c.alignment = Alignment(horizontal=align, vertical="center", wrap_text=True)
        c.font = Font(bold=bold)

    # Encabezados superiores
    merge_set(ws, "A4", "B4", "Tipo de Fondo", True)
    merge_set(ws, "C4", "D4", "Responsable del fondo", True)
    merge_set(ws, "E4", "G4", "Institución", True)
    merge_set(ws, "A5", "B5", st.session_state.data["meta"].get("tipo_fondo",""), False, "left")
    merge_set(ws, "C5", "D5", st.session_state.data["meta"].get("responsable",""), False, "left")
    merge_set(ws, "E5", "G5", st.session_state.data["meta"].get("institucion",""), False, "left")

    merge_set(ws, "A6", "D6", "Fecha Rendición", True)
    merge_set(ws, "E6", "G6", "N° Rendición", True)
    merge_set(ws, "A7", "D7", st.session_state.data["meta"].get("fecha_rendicion",""), False, "left")
    merge_set(ws, "E7", "G7", st.session_state.data["meta"].get("n_rendicion",""), False, "left")

    merge_set(ws, "A9", "D9", "N° REX", True)
    merge_set(ws, "E9", "G9", "Fecha REX", True)
    merge_set(ws, "A10", "D10", st.session_state.data["meta"].get("n_rex",""), False, "left")
    merge_set(ws, "E10", "G10", st.session_state.data["meta"].get("fecha_rex",""), False, "left")
    set_border_range(ws, "A4:G10")
    for r in range(4, 11): ws.row_dimensions[r].height = 18

    # Tabla
    start_row = 12
    for j, h in enumerate(headers, start=1):
        c = ws.cell(row=start_row, column=j, value=h)
        c.font = Font(bold=True); c.alignment = Alignment(horizontal="center"); c.border=border

    df = gastos_df()
    if df.empty:
        for j in range(1, 8):
            c = ws.cell(row=start_row+1, column=j, value="" if j != 1 else "Sin registros")
            c.border = border
            if j==1: c.alignment = Alignment(horizontal="center")
        last_row = start_row + 1
    else:
        for i, row in enumerate(dataframe_to_rows(df, index=False, header=False), start=1):
            for j, val in enumerate(row, start=1):
                c = ws.cell(row=start_row+i, column=j, value=val)
                c.border = border
                if j in (4,5,6):  # N° Doc, Detalle, Proveedor
                    c.alignment = Alignment(horizontal="left", wrap_text=True, vertical="top")
                if j==2 and isinstance(val, date):
                    c.number_format = "yyyy-mm-dd"
                if j==7:
                    c.number_format = '"$"#,##0'; c.alignment = Alignment(horizontal="right")
        last_row = start_row + len(df)

    # Total
    ws.merge_cells(start_row=last_row+1, start_column=1, end_row=last_row+1, end_column=6)
    c = ws.cell(row=last_row+1, column=1, value="Monto Total del Gasto"); c.border=border; c.alignment=Alignment(horizontal="right"); c.font=Font(bold=True)
    c = ws.cell(row=last_row+1, column=7, value=float(df["Monto"].sum() if not df.empty else 0))
    c.border=border; c.alignment=Alignment(horizontal="right"); c.number_format = '"$"#,##0'; c.font = Font(bold=True)

    # ---------- Hoja Resumen ----------
    ws2 = wb.create_sheet("Resumen")
    ws2.page_setup.orientation = "landscape"; ws2.page_setup.fitToWidth = 1
    ws2.page_margins = PageMargins(left=0.3, right=0.3, top=0.5, bottom=0.5)

    if st.session_state.logo_bytes:
        try:
            lf = "/tmp/logo_tmp2.png"
            with open(lf, "wb") as f: f.write(st.session_state.logo_bytes)
            img = XLImage(lf)
            img.width = max(80, min(logo_px, 220))
            img.height = int(img.width * 0.35)
            img.anchor = "A1"
            ws2.add_image(img)
        except Exception:
            pass

    def set_border_range2(rng: str):
        for row in ws2[rng]:
            for c in row:
                c.border = border

    def m(cell1, cell2, txt="", bold=False, align="center"):
        ws2.merge_cells(f"{cell1}:{cell2}")
        c = ws2[cell1]; c.value = txt
        c.font = Font(bold=bold); c.alignment = Alignment(horizontal=align, vertical="center", wrap_text=True)
        return c

    for col, w in zip("ABCDEFGH", [22,28,10,22,10,22,10,12]):
        ws2.column_dimensions[col].width = w

    row = 5
    m("A"+str(row), "B"+str(row), "Tipo de Fondo", True);     m("C"+str(row), "E"+str(row), "Nombre Responsable del Fondo", True); m("F"+str(row), "H"+str(row), "N° RUT", True); row+=1
    m("A"+str(row), "B"+str(row), st.session_state.data["meta"].get("tipo_fondo",""), False, "left")
    m("C"+str(row), "E"+str(row), st.session_state.data["meta"].get("responsable",""), False, "left")
    m("F"+str(row), "H"+str(row), st.session_state.data["meta"].get("rut",""), False, "left"); row+=1

    m("A"+str(row), "B"+str(row), "Institución", True);       m("C"+str(row), "E"+str(row), "Cargo", True);                           m("F"+str(row), "H"+str(row), "N° Rendición", True); row+=1
    m("A"+str(row), "B"+str(row), st.session_state.data["meta"].get("institucion",""), False, "left")
    m("C"+str(row), "E"+str(row), st.session_state.data["meta"].get("cargo",""), False, "left")
    m("F"+str(row), "H"+str(row), st.session_state.data["meta"].get("n_rendicion",""), False, "left"); row+=1

    m("A"+str(row), "B"+str(row), "Mes que Rinde", True);     m("C"+str(row), "E"+str(row), "N° REX", True);                          m("F"+str(row), "H"+str(row), "Fecha REX", True); row+=1
    m("A"+str(row), "B"+str(row), st.session_state.data["meta"].get("mes_que_rinde",""), False, "left")
    m("C"+str(row), "E"+str(row), st.session_state.data["meta"].get("n_rex",""), False, "left")
    m("F"+str(row), "H"+str(row), st.session_state.data["meta"].get("fecha_rex",""), False, "left"); row+=1

    m("A"+str(row), "E"+str(row), "Observaciones", True);     m("F"+str(row), "H"+str(row), "Monto Inicial Fondo", True); row+=1
    m("A"+str(row), "E"+str(row), st.session_state.data["meta"].get("observaciones",""), False, "left")
    c = m("F"+str(row), "H"+str(row), st.session_state.data["fondo_inicial"], False, "right"); c.number_format = '"$"#,##0'; row+=1

    m("A"+str(row), "C"+str(row), "N° Egreso Contable Inicial del Fondo", True); m("D"+str(row), "E"+str(row), "", True)
    m("F"+str(row), "G"+str(row), "Fecha de Egreso Inicial del Fondo", True);    m("H"+str(row), "H"+str(row), "", True); row+=1
    m("A"+str(row), "C"+str(row), st.session_state.data["meta"].get("n_egreso_inicial",""), False, "left")
    m("F"+str(row), "G"+str(row), st.session_state.data["meta"].get("fecha_egreso_inicial",""), False, "left"); row+=2
    set_border_range2(f"A5:H{row-1}")

    m("A"+str(row), "H"+str(row), "CUADRO RESUMEN RENDICION", True); row+=1
    labels_vals = [
        ("Saldo Inicial/Rendición Mes Anterior", parse_float(st.session_state.data["meta"].get("saldo_mes_anterior", 0))),
        ("Monto Recibido Mes anterior", parse_float(st.session_state.data["meta"].get("monto_recibido_mes_anterior", 0))),
        ("Monto Gasto del mes", float(df["Monto"].sum() if not df.empty else 0.0)),
        ("Monto del gasto del mes Transporte", parse_float(st.session_state.data["meta"].get("monto_gasto_transporte", 0))),
    ]
    for label, val in labels_vals:
        m("A"+str(row), "E"+str(row), label, False, "left")
        c = m("F"+str(row), "H"+str(row), val, False, "right"); c.number_format = '"$"#,##0'
        row+=1
    saldo_final = labels_vals[0][1] + labels_vals[1][1] - labels_vals[2][1] - labels_vals[3][1]
    c1 = m("A"+str(row), "E"+str(row), "Saldo Final", True, "left")
    c2 = m("F"+str(row), "H"+str(row), saldo_final, True, "right"); c2.number_format = '"$"#,##0'
    set_border_range2(f"A{row-len(labels_vals)-1}:H{row}")
    row+=3

    def linea_firma(r, c1_, c2_, titulo):
        m(c1_+str(r), c2_+str(r), "_"*40, False, "center")
        m(c1_+str(r+1), c2_+str(r+1), titulo, False, "center")

    linea_firma(row, "B", "D", "Encargado/a del Fondo")
    linea_firma(row, "F", "H", "Director/a Ejecutiva")

    bio = io.BytesIO(); wb.save(bio); return bio.getvalue()

# ---------------------------- UI ----------------------------
init_state()
st.title("Rendición de Fondos Fijos P01 – SLEP Petorca")

with st.sidebar:
    st.header("Configuración general")
    fondo_inicial = st.number_input("Monto inicial del fondo", min_value=0.0, step=1000.0,
                                    value=float(st.session_state.data["fondo_inicial"]))
    if st.button("Guardar fondo"):
        st.session_state.data["fondo_inicial"] = float(fondo_inicial)
        st.success("Fondo inicial actualizado.")

    st.divider()
    st.caption("Logo (se usa en PDF y Excel).")
    logo_file = st.file_uploader("Logo (PNG/JPG)", type=["png","jpg","jpeg"], key="logo_up")
    if logo_file is not None:
        st.session_state.logo_bytes = logo_file.read()
        st.session_state.logo_name = logo_file.name
        st.success(f"Logo cargado: {st.session_state.logo_name}")
    logo_mm = st.slider("Tamaño del logo en PDF (mm)", 16, 40, 24)
    logo_px = st.slider("Tamaño del logo en Excel (px)", 80, 240, 140)

    st.divider()
    st.caption("Firmas (opcional, solo PDF).")
    cols_f = st.columns(2)
    with cols_f[0]:
        f1 = st.file_uploader("Firma Encargado/a", type=["png","jpg","jpeg"], key="f1")
        if f1: st.session_state.firmas["encargado"] = f1.read()
        f3 = st.file_uploader("Firma Revisor/a 1", type=["png","jpg","jpeg"], key="f3")
        if f3: st.session_state.firmas["revisor1"] = f3.read()
        f5 = st.file_uploader("Firma Unidad Finanzas", type=["png","jpg","jpeg"], key="f5")
        if f5: st.session_state.firmas["u_finanzas"] = f5.read()
        f7 = st.file_uploader("Firma Jefe Adm/Finanzas", type=["png","jpg","jpeg"], key="f7")
        if f7: st.session_state.firmas["jefe_adm_fin"] = f7.read()
    with cols_f[1]:
        f2 = st.file_uploader("Firma Director/a", type=["png","jpg","jpeg"], key="f2")
        if f2: st.session_state.firmas["directora"] = f2.read()
        f4 = st.file_uploader("Firma Jefe Unidad", type=["png","jpg","jpeg"], key="f4")
        if f4: st.session_state.firmas["jefe_unidad"] = f4.read()
        f6 = st.file_uploader("Firma Contab. y Finanzas", type=["png","jpg","jpeg"], key="f6")
        if f6: st.session_state.firmas["contab_finanzas"] = f6.read()

    st.divider()
    st.caption("Importar / Exportar datos")
    up = st.file_uploader("Importar datos JSON", type=["json"], key="json_up")
    if up is not None and st.button("Cargar JSON"):
        load_data_from_json(up)
    st.download_button("Exportar datos a JSON", data=export_data_json(),
                       file_name="rendicion_datos.json", mime="application/json")

st.subheader("Metadatos de la rendición")
m = st.session_state.data["meta"]
c1,c2,c3,c4 = st.columns(4)
with c1:
    m["tipo_fondo"] = st.text_input("Tipo de Fondo", value=m.get("tipo_fondo",""))
    m["responsable"] = st.text_input("Responsable del fondo", value=m.get("responsable",""))
    m["rut"] = st.text_input("N° RUT", value=m.get("rut",""))
with c2:
    m["institucion"] = st.text_input("Institución", value=m.get("institucion",""))
    m["cargo"] = st.text_input("Cargo", value=m.get("cargo",""))
    m["mes_que_rinde"] = st.text_input("Mes que rinde", value=m.get("mes_que_rinde",""))
with c3:
    m["fecha_rendicion"] = st.text_input("Fecha Rendición", value=m.get("fecha_rendicion",""))
    m["n_rendicion"] = st.text_input("N° Rendición", value=m.get("n_rendicion",""))
    m["n_rex"] = st.text_input("N° REX", value=m.get("n_rex",""))
with c4:
    m["fecha_rex"] = st.text_input("Fecha REX", value=m.get("fecha_rex",""))
    m["n_egreso_inicial"] = st.text_input("N° Egreso Contable Inicial del Fondo", value=m.get("n_egreso_inicial",""))
    m["fecha_egreso_inicial"] = st.text_input("Fecha de Egreso Inicial del Fondo", value=m.get("fecha_egreso_inicial",""))
m["observaciones"] = st.text_area("Observaciones", value=m.get("observaciones",""))

st.markdown("**Cuadro Resumen (para cálculo del saldo final)**")
c1,c2,c3 = st.columns(3)
with c1:
    m["saldo_mes_anterior"] = st.number_input("Saldo Inicial/Rendición Mes Anterior", value=parse_float(m.get("saldo_mes_anterior",0.0)), step=1000.0)
with c2:
    m["monto_recibido_mes_anterior"] = st.number_input("Monto Recibido Mes anterior", value=parse_float(m.get("monto_recibido_mes_anterior",0.0)), step=1000.0)
with c3:
    m["monto_gasto_transporte"] = st.number_input("Monto del gasto del mes Transporte", value=parse_float(m.get("monto_gasto_transporte",0.0)), step=1000.0)

st.subheader("Registrar gasto")
with st.form("form_gasto", clear_on_submit=True):
    c1,c2,c3,c4,c5,c6 = st.columns([1.1,1.1,1.1,2.4,1.6,1.2])
    with c1: f = st.date_input("Fecha", value=date.today())
    with c2: tipo = st.selectbox("Tipo documento", ["Boleta","Factura","Comprobante","Otro"])
    with c3: ndoc = st.text_input("N° documento")
    with c4: d = st.text_input("Detalle del gasto")
    with c5: prov = st.text_input("Proveedor")
    with c6: mnt = st.number_input("Monto", min_value=0.0, step=1000.0)
    doc = st.file_uploader("Documento (opcional)")
    if st.form_submit_button("Agregar"):
        if d.strip() == "":
            st.error("El detalle es obligatorio.")
        else:
            add_gasto(f, tipo, ndoc, d, prov, mnt, doc); st.success("Gasto agregado.")

st.subheader("Gastos registrados")
df = gastos_df()
if df.empty:
    st.info("Aún no hay gastos.")
else:
    showing = df.copy(); showing["Seleccionar"] = False
    showing["Monto"] = showing["Monto"].apply(money)
    edited = st.data_editor(showing, hide_index=False, use_container_width=True, num_rows="static")
    selected_indices = edited.index[edited["Seleccionar"] == True].tolist()
    if st.button("Eliminar seleccionados"):
        remove_gastos(selected_indices); st.rerun()

st.subheader("Resumen")
fondo, total, saldo, cantidad = totals()
c1,c2,c3,c4 = st.columns(4)
c1.metric("Fondo inicial", money(fondo))
c2.metric("Total gastos", money(total))
c3.metric("Saldo", money(saldo))
c4.metric("Cantidad", f"{cantidad}")

st.subheader("Distribución")
labels = ["Gastos", "Saldo"]
vals = [max(total,0.0), max(saldo,0.0)]
if sum(vals) <= 0:
    st.info("Aún no hay datos para graficar. Configura el fondo inicial o registra gastos.")
else:
    fig, ax = plt.subplots()
    ax.pie(vals, labels=labels, autopct="%1.1f%%"); ax.axis("equal")
    st.pyplot(fig)

st.subheader("Exportaciones")
opt_landscape = st.toggle("Generar PDF en orientación horizontal (recomendado)", value=True)
colx, colp = st.columns(2)
with colx:
    st.download_button("Descargar Excel", data=export_excel(logo_px),
        file_name="rendicion_gastos.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
with colp:
    st.download_button("Descargar PDF", data=export_pdf(opt_landscape, logo_mm),
        file_name="rendicion_gastos.pdf", mime="application/pdf")

st.subheader("Documentos adjuntos")
if len(st.session_state.data["gastos"]) == 0:
    st.caption("No hay documentos adjuntos aún.")
else:
    for i, g in enumerate(st.session_state.data["gastos"]):
        if g.get("bytes_doc"):
            st.download_button(f"Descargar '{g.get('nombre_doc','documento')}'",
                               data=g["bytes_doc"], file_name=g.get("nombre_doc","documento"))
        else:
            if g.get("nombre_doc"):
                st.caption(f"{i+1}. {g.get('nombre_doc')} (no embebido)")
st.caption("⚠️ Nota: Los archivos subidos viven en la sesión. Usa Exportar/Importar JSON para volver a cargar los datos.")
