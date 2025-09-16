
import io
import json
from datetime import date, datetime
from typing import List, Dict, Any, Optional

import streamlit as st
import pandas as pd
from fpdf import FPDF
import matplotlib.pyplot as plt

st.set_page_config(page_title="Rendición de Cuentas – SLEP Petorca", layout="wide")

# ----------------------------
# Helpers & State
# ----------------------------
def init_state():
    if "data" not in st.session_state:
        st.session_state.data = {
            "fondo_inicial": 0.0,
            # gastos: list of dicts: {fecha:str 'YYYY-MM-DD', monto:float, detalle:str, nombre_doc:str or None, bytes_doc: bytes or None}
            "gastos": []
        }
    if "logo_bytes" not in st.session_state:
        st.session_state.logo_bytes = None
        st.session_state.logo_name = None

def money(x: float) -> str:
    try:
        return f"${x:,.0f}".replace(",", ".")
    except Exception:
        return f"${x}"

def load_data_from_json(file) -> None:
    try:
        obj = json.load(file)
        # backward compatibility: allow missing keys
        fi = float(obj.get("fondo_inicial", 0))
        gastos = obj.get("gastos", [])
        fixed_gastos = []
        for g in gastos:
            fixed_gastos.append({
                "fecha": g.get("fecha"),
                "monto": float(g.get("monto", 0)),
                "detalle": g.get("detalle") or g.get("descripcion") or "",
                "nombre_doc": g.get("nombre_doc"),
                "bytes_doc": None  # documents are not embedded when importing JSON (optional)
            })
        st.session_state.data = {"fondo_inicial": fi, "gastos": fixed_gastos}
        st.success("Datos cargados desde JSON.")
    except Exception as e:
        st.error(f"Error al leer JSON: {e}")

def export_data_json() -> bytes:
    # For privacy/size, do not embed raw document bytes in JSON.
    data = {
        "fondo_inicial": st.session_state.data["fondo_inicial"],
        "gastos": [
            {
                "fecha": g["fecha"],
                "monto": g["monto"],
                "detalle": g["detalle"],
                "nombre_doc": g.get("nombre_doc"),
            }
            for g in st.session_state.data["gastos"]
        ]
    }
    return json.dumps(data, ensure_ascii=False, indent=2).encode("utf-8")

def gastos_df() -> pd.DataFrame:
    df = pd.DataFrame(st.session_state.data["gastos"])
    if not df.empty:
        df = df.assign(
            Fecha=pd.to_datetime(df["fecha"]).dt.date,
            Detalle=df["detalle"],
            Monto=df["monto"].astype(float),
            Documento=df["nombre_doc"].fillna("—")
        )[["Fecha", "Detalle", "Monto", "Documento"]]
    return df

def totals():
    df = gastos_df()
    total = float(df["Monto"].sum()) if not df.empty else 0.0
    cantidad = int(len(df)) if not df.empty else 0
    fondo = float(st.session_state.data["fondo_inicial"])
    saldo = fondo - total
    return fondo, total, saldo, cantidad

def add_gasto(fecha: date, detalle: str, monto: float, doc_file):
    nombre_doc = None
    bytes_doc = None
    if doc_file is not None:
        nombre_doc = doc_file.name
        bytes_doc = doc_file.read()
    st.session_state.data["gastos"].append({
        "fecha": fecha.strftime("%Y-%m-%d"),
        "monto": float(monto),
        "detalle": detalle,
        "nombre_doc": nombre_doc,
        "bytes_doc": bytes_doc
    })

def remove_gastos(indices: List[int]):
    # Remove from last to first to keep indices stable
    for idx in sorted(indices, reverse=True):
        if 0 <= idx < len(st.session_state.data["gastos"]):
            st.session_state.data["gastos"].pop(idx)

def export_excel() -> bytes:
    df = gastos_df()
    from openpyxl import Workbook
    from openpyxl.utils.dataframe import dataframe_to_rows
    from openpyxl.styles import Alignment, Font
    from openpyxl.worksheet.table import Table, TableStyleInfo

    wb = Workbook()
    ws = wb.active
    ws.title = "Gastos"

    # Title
    ws["A1"] = "Rendición de Cuentas – SLEP Petorca"
    ws["A1"].font = Font(bold=True, size=14)
    ws.merge_cells("A1:D1")
    ws["A1"].alignment = Alignment(horizontal="center")

    # Table
    if df.empty:
        df = pd.DataFrame(columns=["Fecha","Detalle","Monto","Documento"])
    start_row = 3
    for r in dataframe_to_rows(df, index=False, header=True):
        ws.append(r)
    # Convert appended data starting at row 3
    last_row = start_row + len(df)
    table = Table(displayName="TablaGastos", ref=f"A{start_row}:D{last_row}")
    style = TableStyleInfo(name="TableStyleMedium2", showFirstColumn=False,
                           showLastColumn=False, showRowStripes=True, showColumnStripes=False)
    table.tableStyleInfo = style
    ws.add_table(table)

    # Summary
    fondo, total, saldo, cantidad = totals()
    ws[f"A{last_row+2}"] = "Fondo inicial"
    ws[f"B{last_row+2}"] = fondo
    ws[f"A{last_row+3}"] = "Total gastos"
    ws[f"B{last_row+3}"] = total
    ws[f"A{last_row+4}"] = "Saldo"
    ws[f"B{last_row+4}"] = saldo
    ws[f"A{last_row+5}"] = "Cantidad de gastos"
    ws[f"B{last_row+5}"] = cantidad

    bio = io.BytesIO()
    wb.save(bio)
    return bio.getvalue()

def export_pdf() -> bytes:
    df = gastos_df()
    fondo, total, saldo, cantidad = totals()

    pdf = FPDF(orientation="P", unit="mm", format="A4")
    pdf.add_page()
    # Logo (optional)
    if st.session_state.logo_bytes:
        try:
            logo_bio = io.BytesIO(st.session_state.logo_bytes)
            pdf.image(logo_bio, x=10, y=8, w=30)
        except Exception:
            pass
    pdf.set_font("Arial", size=12)
    pdf.cell(0, 10, "Rendición de Cuentas – SLEP Petorca", ln=True, align="C")
    pdf.ln(5)
    pdf.set_font("Arial", size=10)
    pdf.cell(0, 6, f"Fecha de emisión: {datetime.now().strftime('%Y-%m-%d %H:%M')}", ln=True)
    pdf.ln(2)

    # Summary
    pdf.set_font("Arial", "B", 11)
    pdf.cell(0, 7, "Resumen", ln=True)
    pdf.set_font("Arial", size=10)
    pdf.cell(0, 6, f"Fondo inicial: {money(fondo)}", ln=True)
    pdf.cell(0, 6, f"Total gastos: {money(total)}", ln=True)
    pdf.cell(0, 6, f"Saldo: {money(saldo)}", ln=True)
    pdf.cell(0, 6, f"Cantidad de gastos: {cantidad}", ln=True)
    pdf.ln(4)

    # Table header
    pdf.set_font("Arial", "B", 10)
    pdf.cell(30, 7, "Fecha", border=1, align="C")
    pdf.cell(100, 7, "Detalle", border=1, align="C")
    pdf.cell(30, 7, "Monto", border=1, align="C")
    pdf.cell(30, 7, "Documento", border=1, align="C")
    pdf.ln()

    pdf.set_font("Arial", size=9)
    if df.empty:
        pdf.cell(190, 7, "Sin registros", border=1, align="C")
        pdf.ln()
    else:
        for _, row in df.iterrows():
            pdf.cell(30, 7, row["Fecha"].strftime("%Y-%m-%d"), border=1)
            # detalle: truncate to fit
            detalle = str(row["Detalle"])[:60]
            pdf.cell(100, 7, detalle, border=1)
            pdf.cell(30, 7, money(float(row["Monto"])), border=1, align="R")
            pdf.cell(30, 7, str(row["Documento"])[:14], border=1)
            pdf.ln()

    bio = io.BytesIO()
    pdf.output(bio)
    return bio.getvalue()

# ----------------------------
# UI
# ----------------------------
init_state()
st.title("Rendición de Cuentas – SLEP Petorca")

with st.sidebar:
    st.header("Configuración")
    fondo_inicial = st.number_input("Fondo inicial", min_value=0.0, step=1000.0, value=float(st.session_state.data["fondo_inicial"]))
    if st.button("Guardar fondo"):
        st.session_state.data["fondo_inicial"] = float(fondo_inicial)
        st.success("Fondo inicial actualizado.")

    st.divider()
    st.caption("Logo opcional para PDF")
    logo_file = st.file_uploader("Logo (PNG/JPG)", type=["png", "jpg", "jpeg"], key="logo_up")
    if logo_file is not None:
        st.session_state.logo_bytes = logo_file.read()
        st.session_state.logo_name = logo_file.name
        st.success(f"Logo cargado: {st.session_state.logo_name}")

    st.divider()
    st.caption("Importar / Exportar datos")
    up = st.file_uploader("Importar datos JSON", type=["json"], key="json_up")
    if up is not None and st.button("Cargar JSON"):
        load_data_from_json(up)
    st.download_button("Exportar datos a JSON", data=export_data_json(), file_name="rendicion_datos.json", mime="application/json")

# ---- Registro de gasto ----
st.subheader("Registrar gasto")
with st.form("form_gasto", clear_on_submit=True):
    cols = st.columns([1, 3, 1.2, 1.6])
    with cols[0]:
        f = st.date_input("Fecha", value=date.today())
    with cols[1]:
        d = st.text_input("Detalle")
    with cols[2]:
        m = st.number_input("Monto", min_value=0.0, step=1000.0)
    with cols[3]:
        doc = st.file_uploader("Documento (opcional)", type=None)
    submitted = st.form_submit_button("Agregar")
    if submitted:
        if d.strip() == "":
            st.error("El detalle es obligatorio.")
        else:
            add_gasto(f, d, m, doc)
            st.success("Gasto agregado.")

# ---- Tabla / Gestión ----
st.subheader("Gastos registrados")
df = gastos_df()
if df.empty:
    st.info("Aún no hay gastos.")
else:
    # Show table with a selection column for deletion
    df_show = df.copy()
    df_show["Seleccionar"] = False
    edited = st.data_editor(df_show, hide_index=False, use_container_width=True, num_rows="static")
    # gather selected rows by index in the edited df
    selected_indices = edited.index[edited["Seleccionar"] == True].tolist()
    # Map back to session_state list indices (same ordering)
    if st.button("Eliminar seleccionados"):
        remove_gastos(selected_indices)
        st.rerun()

# ---- Resumen ----
st.subheader("Resumen")
fondo, total, saldo, cantidad = totals()
c1, c2, c3, c4 = st.columns(4)
c1.metric("Fondo inicial", money(fondo))
c2.metric("Total gastos", money(total))
c3.metric("Saldo", money(saldo))
c4.metric("Cantidad", f"{cantidad}")

# ---- Gráfico ----
st.subheader("Distribución")
labels = ["Gastos", "Saldo"]
vals = [max(total, 0), max(saldo, 0)]
fig, ax = plt.subplots()
ax.pie(vals, labels=labels, autopct="%1.1f%%")
ax.axis("equal")
st.pyplot(fig)

# ---- Exportaciones ----
st.subheader("Exportaciones")
colx, colp = st.columns(2)
with colx:
    st.download_button("Descargar Excel", data=export_excel(), file_name="rendicion_gastos.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
with colp:
    st.download_button("Descargar PDF", data=export_pdf(), file_name="rendicion_gastos.pdf", mime="application/pdf")

# ---- Descarga de documentos ----
st.subheader("Documentos adjuntos")
if len(st.session_state.data["gastos"]) == 0:
    st.caption("No hay documentos adjuntos aún.")
else:
    for i, g in enumerate(st.session_state.data["gastos"]):
        if g.get("bytes_doc"):
            st.download_button(f"Descargar '{g.get('nombre_doc','documento')}'", data=g["bytes_doc"], file_name=g.get("nombre_doc","documento"))
        else:
            if g.get("nombre_doc"):
                st.caption(f"{i+1}. {g.get('nombre_doc')} (no embebido)")

st.caption("⚠️ Nota: Los archivos subidos viven en la sesión del navegador (memoria). Exporta/Importa JSON para persistir registros entre sesiones.")
