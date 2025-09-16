from datetime import date, datetime
import calendar
import json
from pathlib import Path
import io

try:
    from kivy.app import App
    from kivy.uix.boxlayout import BoxLayout
    from kivy.uix.label import Label
    from kivy.uix.textinput import TextInput
    from kivy.uix.button import Button
    from kivy.uix.popup import Popup
    from kivy.uix.filechooser import FileChooserIconView
    from kivy.uix.scrollview import ScrollView
    from kivy.uix.gridlayout import GridLayout
    from kivy.uix.checkbox import CheckBox
    from kivy.uix.image import Image
except ModuleNotFoundError as exc:  # pragma: no cover - Kivy missing
    raise SystemExit(
        "Kivy no esta instalado. Ejecute 'python -m pip install kivy'"
    ) from exc


DATA_FILE = Path("gastos.json")
# Ruta del logotipo a incluir en el PDF
LOGO_PATH = r"C:\PY\Logotipo Petorca-01.png"


class GastosManager:
    """Gestiona los datos de gastos y el fondo inicial."""

    def __init__(self, data_file=DATA_FILE):
        self.data_file = data_file
        self.data = {"fondo_inicial": 0, "gastos": []}
        self.cargar_datos()

    def cargar_datos(self):
        if self.data_file.exists():
            with self.data_file.open("r", encoding="utf-8") as f:
                self.data = json.load(f)

    def guardar_datos(self):
        with self.data_file.open("w", encoding="utf-8") as f:
            json.dump(self.data, f, ensure_ascii=False, indent=2)

    def establecer_fondo(self, monto):
        self.data["fondo_inicial"] = monto
        self.guardar_datos()

    def agregar_gasto(self, monto, descripcion, documento, fecha=None):
        gasto = {
            "fecha": (fecha or date.today()).isoformat(),
            "monto": monto,
            "descripcion": descripcion,
            "documento": documento,
        }
        self.data.setdefault("gastos", []).append(gasto)
        self.guardar_datos()

    def resumen(self):
        total = sum(g["monto"] for g in self.data.get("gastos", []))
        cantidad = len(self.data.get("gastos", []))
        saldo = self.data.get("fondo_inicial", 0) - total
        return cantidad, total, saldo

    def exportar_excel(self, path="informe_gastos.xlsx"):
        """Genera un archivo Excel con todos los gastos."""
        try:  # Importar solo cuando se usa
            from openpyxl import Workbook
        except Exception as exc:  # pragma: no cover - dependencia faltante
            raise RuntimeError(
                "Debe instalar openpyxl para exportar a Excel"
            ) from exc

        wb = Workbook()
        ws = wb.active
        ws.title = "Gastos"

        # Titulo
        ws.append([
            "Rendicion de Gastos menores Fondo Fijo - SLEP Petorca - Programa 01"
        ])
        ws.append([])

        ws.append(["Fecha", "Detalle", "Monto", "Documento"])
        gastos = sorted(self.data.get("gastos", []), key=lambda g: g["fecha"])
        for g in gastos:
            ws.append([
                g["fecha"],
                g["descripcion"],
                g["monto"],
                Path(g["documento"]).name if g["documento"] else "",
            ])

        ws.append([])
        cantidad, total, saldo = self.resumen()
        ws.append(["", "Fondo entregado", self.data.get("fondo_inicial", 0)])
        ws.append(["", "Gastos realizados", total])
        ws.append(["", "Saldo disponible", saldo])

        wb.save(path)

    def exportar_pdf(self, path="informe_gastos.pdf"):
        """Genera un PDF con el detalle de gastos."""
        try:
            from fpdf import FPDF
        except Exception as exc:  # pragma: no cover - dependencia faltante
            raise RuntimeError(
                "Debe instalar fpdf para exportar a PDF"
            ) from exc

        # Crear en orientacion horizontal
        pdf = FPDF(orientation="L")
        pdf.add_page()
        # Incluir el logotipo institucional si esta disponible
        if Path(LOGO_PATH).exists():
            pdf.image(LOGO_PATH, x=10, y=8, w=40)
        pdf.set_font("Helvetica", "B", 12)
        pdf.cell(0, 10,
                 "Rendicion de Gastos menores Fondo Fijo - SLEP Petorca - Programa 01",
                 ln=True, align="C")
        pdf.ln(5)

        pdf.set_font("Helvetica", size=10)
        fecha_w = 30
        detalle_w = 120
        monto_w = 30
        # Ancho restante para la columna Documento
        max_doc_w = (
            pdf.w
            - pdf.l_margin
            - pdf.r_margin
            - fecha_w
            - detalle_w
            - monto_w
        )
        gastos = sorted(self.data.get("gastos", []), key=lambda g: g["fecha"])
        docs = [Path(g["documento"]).name if g["documento"] else "" for g in gastos]
        if docs:
            doc_w = min(max(pdf.get_string_width(d) + 4 for d in docs), max_doc_w)
        else:
            doc_w = max_doc_w
        pdf.cell(fecha_w, 8, "Fecha", border=1)
        pdf.cell(detalle_w, 8, "Detalle", border=1)
        pdf.cell(monto_w, 8, "Monto", border=1)
        pdf.cell(doc_w, 8, "Documento", border=1, ln=True)

        def wrap_text(text, width):
            words = text.split()
            lines = []
            line = ""
            for word in words:
                test = f"{line} {word}".strip()
                if pdf.get_string_width(test) <= width:
                    line = test
                else:
                    lines.append(line)
                    line = word
            if line:
                lines.append(line)
            return lines or [""]

        line_h = 8
        for g in gastos:
            doc_text = Path(g["documento"]).name if g["documento"] else ""
            doc_lines = wrap_text(doc_text, doc_w - 2)
            row_h = line_h * max(1, len(doc_lines))
            x_left = pdf.l_margin
            y_start = pdf.get_y()
            pdf.cell(fecha_w, row_h, g["fecha"], border=1)
            pdf.cell(detalle_w, row_h, g["descripcion"], border=1)
            pdf.cell(monto_w, row_h, f"${g['monto']}", border=1)
            pdf.multi_cell(doc_w, line_h, "\n".join(doc_lines), border=1)
            y_end = max(y_start + row_h, pdf.get_y())
            pdf.set_xy(x_left, y_end)

        pdf.ln(5)
        cantidad, total, saldo = self.resumen()

        pdf.cell(0, 6, "Resumen", ln=True)
        pdf.cell(60, 8, "Total Fondo Entregado", border=1)
        pdf.cell(40, 8, str(self.data.get("fondo_inicial", 0)), border=1, ln=True)
        pdf.cell(60, 8, "Total Gastos Realizados", border=1)
        pdf.cell(40, 8, str(total), border=1, ln=True)
        pdf.cell(60, 8, "Saldo disponible", border=1)
        pdf.cell(40, 8, str(saldo), border=1, ln=True)

        pdf.ln(15)
        pdf.cell(0, 6, "_____________________________", ln=True, align="L")
        pdf.cell(0, 6, "Nombre:", ln=True)
        pdf.cell(0, 6, "RUT:", ln=True)
        pdf.cell(0, 6, "Cargo:", ln=True)

        pdf.output(path)


class FileChooserPopup(Popup):
    def __init__(self, on_select, **kwargs):
        super().__init__(title="Seleccionar documento", size_hint=(0.9, 0.9), **kwargs)
        self.on_select = on_select
        box = BoxLayout(orientation="vertical")
        self.fc = FileChooserIconView()
        box.add_widget(self.fc)
        btns = BoxLayout(size_hint_y=None, height="40dp")
        ok = Button(text="Aceptar")
        ok.bind(on_press=self._ok)
        cancel = Button(text="Cancelar")
        cancel.bind(on_press=lambda *_: self.dismiss())
        btns.add_widget(ok)
        btns.add_widget(cancel)
        box.add_widget(btns)
        self.add_widget(box)

    def _ok(self, _instance):
        if self.fc.selection:
            self.on_select(self.fc.selection[0])
        self.dismiss()


class CalendarPopup(Popup):
    """Popup para elegir una fecha navegando por meses y anios."""

    def __init__(self, on_select, year=None, month=None, **kwargs):
        self.year = year or datetime.now().year
        self.month = month or datetime.now().month
        super().__init__(title="", size_hint=(0.8, 0.8), **kwargs)
        self.on_select = on_select

        self.box = BoxLayout(orientation="vertical")
        self.add_widget(self.box)

        # Barra de navegacion
        nav = BoxLayout(size_hint_y=None, height=40)
        btn_prev_year = Button(text="<<")
        btn_prev_year.bind(on_press=self._prev_year)
        btn_prev_month = Button(text="<")
        btn_prev_month.bind(on_press=self._prev_month)
        self.lbl_title = Label(size_hint_x=2)
        btn_next_month = Button(text=">")
        btn_next_month.bind(on_press=self._next_month)
        btn_next_year = Button(text=">>")
        btn_next_year.bind(on_press=self._next_year)
        nav.add_widget(btn_prev_year)
        nav.add_widget(btn_prev_month)
        nav.add_widget(self.lbl_title)
        nav.add_widget(btn_next_month)
        nav.add_widget(btn_next_year)
        self.box.add_widget(nav)

        # Contenedor para los dias
        self.scroll = ScrollView()
        self.box.add_widget(self.scroll)

        self._build_calendar()

    def _build_calendar(self):
        self.title = f"{calendar.month_name[self.month]} {self.year}"
        self.lbl_title.text = self.title
        grid = GridLayout(cols=7, spacing=2, padding=2, size_hint_y=None)
        grid.bind(minimum_height=grid.setter("height"))
        for name in ["Lu", "Ma", "Mi", "Ju", "Vi", "Sa", "Do"]:
            grid.add_widget(Label(text=name, size_hint_y=None, height=30))
        for week in calendar.monthcalendar(self.year, self.month):
            for day in week:
                if day == 0:
                    grid.add_widget(Label(text="", size_hint_y=None, height=30))
                else:
                    btn = Button(text=str(day), size_hint_y=None, height=30)
                    btn.bind(on_press=lambda inst, d=day: self._choose(d))
                    grid.add_widget(btn)
        self.scroll.clear_widgets()
        self.scroll.add_widget(grid)

    def _prev_month(self, _instance):
        self.month -= 1
        if self.month < 1:
            self.month = 12
            self.year -= 1
        self._build_calendar()

    def _next_month(self, _instance):
        self.month += 1
        if self.month > 12:
            self.month = 1
            self.year += 1
        self._build_calendar()

    def _prev_year(self, _instance):
        self.year -= 1
        self._build_calendar()

    def _next_year(self, _instance):
        self.year += 1
        self._build_calendar()

    def _choose(self, day):
        self.on_select(date(self.year, self.month, day))
        self.dismiss()


class GastosUI(BoxLayout):
    """Interfaz principal de la aplicacion."""

    def __init__(self, **kwargs):
        super().__init__(orientation="horizontal", spacing=10, padding=10, **kwargs)
        self.manager = GastosManager()
        self.selected_indices = set()
        self.edit_index = None

        self.controls_box = BoxLayout(orientation="vertical", spacing=5, size_hint_x=0.5)
        self.add_widget(self.controls_box)

        self.lista_layout = GridLayout(cols=1, spacing=5, size_hint_y=None)
        self.lista_layout.bind(minimum_height=self.lista_layout.setter("height"))

        self.right_box = BoxLayout(orientation="vertical", spacing=5, size_hint_x=0.5)
        scroll = ScrollView()
        scroll.add_widget(self.lista_layout)
        self.chart_image = Image(size_hint_y=None, height=200)
        self.right_box.add_widget(scroll)
        self.right_box.add_widget(self.chart_image)
        self.add_widget(self.right_box)
        self.actualizar_lista()

        # Fondo inicial
        self.fondo_input = TextInput(
            text=str(self.manager.data.get("fondo_inicial", 0)),
            hint_text="Fondo inicial",
            input_filter="float",
        )
        self.controls_box.add_widget(self.fondo_input)
        btn_fondo = Button(text="Establecer fondo")
        btn_fondo.bind(on_press=self.on_set_fondo)
        self.controls_box.add_widget(btn_fondo)

        # Registro de gasto
        self.detalle_input = TextInput(hint_text="Detalle del gasto")
        self.controls_box.add_widget(self.detalle_input)
        self.monto_input = TextInput(hint_text="Monto", input_filter="float")
        self.controls_box.add_widget(self.monto_input)
        self.doc_label = Label(text="Documento: (ninguno)")
        self.controls_box.add_widget(self.doc_label)
        btn_doc = Button(text="Seleccionar documento")
        btn_doc.bind(on_press=self.on_select_doc)
        self.controls_box.add_widget(btn_doc)
        self.doc_path = ""
        self.fecha = date.today()
        self.fecha_label = Label(text=f"Fecha: {self.fecha.isoformat()}")
        self.controls_box.add_widget(self.fecha_label)
        btn_fecha = Button(text="Elegir fecha")
        btn_fecha.bind(on_press=self.on_fecha)
        self.controls_box.add_widget(btn_fecha)
        self.btn_registrar = Button(text="Registrar gasto")
        self.btn_registrar.bind(on_press=self.on_registrar)
        self.controls_box.add_widget(self.btn_registrar)
        btn_edit = Button(text="Editar seleccionado")
        btn_edit.bind(on_press=self.on_editar)
        self.controls_box.add_widget(btn_edit)
        btn_delete = Button(text="Eliminar seleccionados")
        btn_delete.bind(on_press=self.on_eliminar)
        self.controls_box.add_widget(btn_delete)

        btn_resumen = Button(text="Mostrar resumen")
        btn_resumen.bind(on_press=self.on_resumen)
        self.controls_box.add_widget(btn_resumen)
        btn_xls = Button(text="Descargar Excel")
        btn_xls.bind(on_press=self.on_export_excel)
        self.controls_box.add_widget(btn_xls)
        btn_pdf = Button(text="Descargar PDF")
        btn_pdf.bind(on_press=self.on_export_pdf)
        self.controls_box.add_widget(btn_pdf)
        self.label_resumen = Label(text="")
        self.controls_box.add_widget(self.label_resumen)

    def actualizar_grafico(self):
        try:
            import matplotlib.pyplot as plt
            from kivy.core.image import Image as CoreImage
        except Exception:
            self.chart_image.texture = None
            return

        fondo = self.manager.data.get("fondo_inicial", 0)
        total = sum(g["monto"] for g in self.manager.data.get("gastos", []))
        saldo = fondo - total
        labels = ["Gastos realizados", "Saldo disponible"]
        sizes = [total, max(saldo, 0)]
        fig, ax = plt.subplots(figsize=(3, 3))
        ax.pie(sizes, labels=labels, autopct="%1.1f%%")
        ax.axis("equal")
        buf = io.BytesIO()
        fig.savefig(buf, format="png")
        plt.close(fig)
        buf.seek(0)
        img = CoreImage(buf, ext="png")
        self.chart_image.texture = img.texture

    def on_set_fondo(self, _instance):
        try:
            fondo = float(self.fondo_input.text)
        except ValueError:
            self.label_resumen.text = "Monto inicial invalido"
            return
        self.manager.establecer_fondo(fondo)
        self.label_resumen.text = "Fondo actualizado"
        self.actualizar_grafico()

    def on_registrar(self, _instance):
        try:
            monto = float(self.monto_input.text)
        except ValueError:
            self.label_resumen.text = "Monto invalido"
            return
        detalle = self.detalle_input.text
        doc = self.doc_path
        if self.edit_index is None:
            self.manager.agregar_gasto(monto, detalle, doc, self.fecha)
            self.label_resumen.text = "Gasto registrado"
        else:
            self.manager.data["gastos"][self.edit_index] = {
                "fecha": self.fecha.isoformat(), 
                "monto": monto,
                "descripcion": detalle,
                "documento": doc,
            }
            self.manager.guardar_datos()
            self.label_resumen.text = "Gasto actualizado"
            self.edit_index = None
            self.btn_registrar.text = "Registrar gasto"
        self.detalle_input.text = ""
        self.monto_input.text = ""
        self.doc_path = ""
        self.doc_label.text = "Documento: (ninguno)"
        self.selected_indices.clear()
        self.actualizar_lista()

    def on_resumen(self, _instance):
        cantidad, total, saldo = self.manager.resumen()
        self.label_resumen.text = (
            f"Gastos: {cantidad} | Total: ${total} | Saldo disponible: ${saldo}"
        )

    def on_export_excel(self, _instance):
        try:
            self.manager.exportar_excel()
            self.label_resumen.text = "Informe Excel creado"
        except Exception as exc:
            self.label_resumen.text = str(exc)

    def on_export_pdf(self, _instance):
        try:
            self.manager.exportar_pdf()
            self.label_resumen.text = "Informe PDF creado"
        except Exception as exc:
            self.label_resumen.text = str(exc)

    def on_select_doc(self, _instance):
        FileChooserPopup(self.set_doc).open()

    def set_doc(self, path):
        self.doc_path = path
        self.doc_label.text = f"Documento: {Path(path).name}"

    def on_fecha(self, _instance):
        CalendarPopup(self.set_fecha).open()

    def set_fecha(self, fecha):
        self.fecha = fecha
        self.fecha_label.text = f"Fecha: {self.fecha.isoformat()}"

    def on_editar(self, _instance):
        if len(self.selected_indices) != 1:
            self.label_resumen.text = "Seleccione un solo registro"
            return
        idx = next(iter(self.selected_indices))
        gasto = self.manager.data["gastos"][idx]
        self.detalle_input.text = gasto["descripcion"]
        self.monto_input.text = str(gasto["monto"])
        self.doc_path = gasto["documento"]
        nombre = Path(gasto["documento"]).name if gasto["documento"] else "(ninguno)"
        self.doc_label.text = f"Documento: {nombre}"
        self.fecha = date.fromisoformat(gasto["fecha"])
        self.fecha_label.text = f"Fecha: {self.fecha.isoformat()}"
        self.edit_index = idx
        self.btn_registrar.text = "Actualizar gasto"
        self.label_resumen.text = "Modifique los campos y confirme"

    def on_eliminar(self, _instance):
        if not self.selected_indices:
            self.label_resumen.text = "No hay registros seleccionados"
            return
        for idx in sorted(self.selected_indices, reverse=True):
            del self.manager.data["gastos"][idx]
        self.manager.guardar_datos()
        self.selected_indices.clear()
        self.edit_index = None
        self.btn_registrar.text = "Registrar gasto"
        self.actualizar_lista()
        self.label_resumen.text = "Registro(s) eliminado(s)"

    def actualizar_lista(self):
        self.lista_layout.clear_widgets()
        for idx, g in enumerate(self.manager.data.get("gastos", [])):
            info = (
                f"{g['fecha']} - {g['descripcion']} - ${g['monto']} - "
                f"{Path(g['documento']).name if g['documento'] else ''}"
            )
            row = BoxLayout(size_hint_y=None, height=30)
            cb = CheckBox(size_hint_x=None, width=30)
            cb.active = idx in self.selected_indices
            cb.bind(active=lambda inst, val, i=idx: self._toggle(i, val))
            row.add_widget(cb)
            row.add_widget(Label(text=info))
            self.lista_layout.add_widget(row)

        self.actualizar_grafico()

    def _toggle(self, idx, active):
        if active:
            self.selected_indices.add(idx)
        else:
            self.selected_indices.discard(idx)


class GastosApp(App):
    def build(self):
        self.title = "App Registro Rendición de Cuentas - SLEP Petorca"
        return GastosUI()


if __name__ == "__main__":
    GastosApp().run()